import datetime
import json
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from reader.read_template import read_template, str_to_date, template_to_mask
from reader import available_sheets_name

def find_order_prot_data(path: str, lab: str, template_name: str):
    template = read_template(lab, template_name)
    wb = load_workbook(path)
    
    for ws in wb.worksheets:
        if ws.title not in available_sheets_name: continue

        prot_num = read_rows(ws, template['PROT']['NUMBER'])
        prot_date = read_rows(ws, template['PROT']['DATE'])
        order_num = read_rows(ws, template['ORDER']['NUMBER'])
        order_date = read_rows(ws, template['ORDER']['DATE'])
    
    
    prot_date = str_to_date(prot_date)
    order_date = str_to_date(order_date)
    wb.close()
    return prot_num, prot_date, order_num, order_date

def read_rows(ws: Worksheet, template: json):
    
    for row in range(1, ws.max_row):
        words: list = []

        for column in range(1, ws.max_column):
            val = ws.cell(row, column).value

            if val is None: continue

            if isinstance(val, datetime.datetime): val = val.strftime("%m/%d/%Y")
        
            words.append(str(val))

        s_row = ' '.join([w for w in words])
        word = check(s_row, template)
        if (word != None):
            return word
    
    return None

def check(s_row: str, template: json):
    # Ищем ключевое слово в данной строке 
    searches = template['SEARCH']
    s_in = False
    for s in searches:
        if s in s_row:
            s_in = True

    if s_in == False: return

    # Иногда по ключевому слову можно найти совсем другую строку,
    # в таких случаях нужно составить словарь ключей которые входят
    # в неправильную строку, чтобы можно было исключать их
    wrongs = template['WRONG']
    for w in wrongs:
        if w in s_row: return

    after = template['AFTER']
    after_index = -1
    words = s_row.split()
    
    # Ищем индекс символа посе которого идет нужное слово
    for i in range(len(words)):
        if words[i] in after:
            after_index = i
            break

    if after_index == -1: return
    words = words[after_index + 1:]

    for _template in template['TEMPLATE']:
        _template = template_to_mask(_template)

        for i in range(len(words)):
            _word = ' '.join([w for w in words[0:i+1]])
            _word = template_to_mask(_word)
            if _template == _word:
                _word = ' '.join([w for w in words[0:i+1]])
                return _word

    return None