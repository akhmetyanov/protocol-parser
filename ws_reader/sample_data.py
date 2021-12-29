from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from reader.read_template import read_template

def read(addresses: list, path: str):
    wb = load_workbook(path)
    
    datas = []

    for element in addresses:
        _addresses = element['addresses']
        _element = element['element']
        _value_column = element['value_column']

        for address in _addresses:
            _row = address['row']
            _col = address['col']
            _ws_name = address['ws_name']

            ws = wb[_ws_name]
            
            ns = str(ws.cell(_row, _col).value)
            value = str(ws.cell(_row, _value_column).value)

            datas.append({
                'element': _element,
                'ns': ns,
                'value': value,
            })

    return datas

def rules(datas: list, lab: str, template_name: str):
    
    template = read_template(lab, template_name)
    operations = {
        'devide':devide,
        'replace':replace
    }

    for i in range(len(datas)):
        name = datas[i]['element']
        value = datas[i]['value']

        # if value == '**':
        #     print(value)

        for r in template['ELEMENTS'][name]['RULES']:
            value2 = rule_to_value(r, value, operations)
            if value2 is not None:
                datas[i]['value2'] = value2
                datas[i]['exlcude'] = r['exlcude']
            
        if 'value2' not in datas[i].keys():
            datas[i]['value2'] = datas[i]['value']
            datas[i]['exlcude'] = 0
            



def rule_to_value(rule: dict, value: str, operations: dict):
    _value = None
    if rule['compare'] == 'full':
        
        if value.strip() == rule['value']:
            f = operations[rule['operation']]
            _value = f(value,  rule['replaceValue'])

    elif rule['compare'] == 'include':

        if  rule['value'] in value:
            f = operations[rule['operation']]
            _value = f(value,  rule['replaceValue'])

    return _value

def devide(value: str, by: str):
    value = value.replace(',','.')
    value = int(value)
    by = int(by)
    value = value/by
    return value

def replace(old: str, new: str):
    old = old.replace(old, new)
    return old