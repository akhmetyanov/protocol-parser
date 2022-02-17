import json
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import true
from reader import available_sheets_name
from reader.read_template import read_template, template_to_mask

def find_element_sample_adress(path: str, lab: str, template_name: str):
    # Поиск в протколе заголовка колонны замеров по элементу
    wb = load_workbook(path, data_only=True)
    template = read_template(lab, template_name)
    for ws in wb.worksheets:
        if ws.title not in available_sheets_name: continue
        element_addresses = finding(ws, template)
        if element_addresses is not None:
            return element_addresses
    wb.close()
    return None

def finding(ws: Worksheet, template: json):
    elements = template['ELEMENTS']
    element_addresses = []
    for el in elements:
        el_template = elements[el]
        header_addresses = find_header(ws, el_template)

        if header_addresses is not None:

            for address in header_addresses:
                row = address['row']
                col = address['col']
                addresses = find_samples(ws, row, col, template['SAMPLE'])
                element_addresses.append({
                    'element': el,
                    'addresses': addresses,
                    'value_column': col
                })

    if len(element_addresses) == 0: return None

    return element_addresses

def find_header(ws: Worksheet, template: json, r: int = 1, c: int = 1):
    header_addresses = []
    for header in template['HEADER']:
        header = str(header).strip().upper()
        for row in range(r, ws.max_row):

            for col in range(c, ws.max_column):

                if ws.cell(row, col).value is None: continue
                value = str(ws.cell(row, col).value).strip().upper()

                # Тут может быть два случаея, когда заголовок имеет подзаголовок и нет.
                # В слуае когда есть подзаголвок, то нужно рекурсивно найти последний подзаголвоок

                if header == value and template['SUBHEADER'] != '0':
                    _header_addresses = find_header(ws, template['SUBHEADER'], row, col)
                    header_addresses.append({
                        'row': _header_addresses[0]['row'],
                        'col': _header_addresses[0]['col']
                    })

                elif header == value and template['SUBHEADER'] == '0':
                    header_addresses.append({
                        'row': row,
                        'col': col
                    })

    return header_addresses

def find_samples(ws: Worksheet, r: int, c:int, template: json):
    for i in range(len(template['TEMPLATE'])):
        template['TEMPLATE'][i] = template_to_mask(template['TEMPLATE'][i])

    # нужно найти номер столбца номеров проб
    _column = -1
    _row = -1
    _flag = False
    for row in range(r, ws.max_row):
        for col in range(c, 0, -1):
            value = ws.cell(row, col).value
            if value is None: continue
            value = str(value).strip().upper()
            value = template_to_mask(value)
            
            for samp_temp in template['TEMPLATE']:
                if str(samp_temp).upper() == value:
                    _column = col
                    _row = row
                    _flag = True
                    break
            
            if _flag: break

        if _flag: break

    if _column == -1: return None

    # читать пока в столбце номера проб соответсвуют шаблону
    finded_address = []
    _flag = True
    while(_flag):
        value = str(ws.cell(_row, _column).value).strip().upper()
        value = template_to_mask(value)

        _not_found = True
        for samp_temp in template['TEMPLATE']:
            if str(samp_temp).upper() == value:
                finded_address.append({
                    'row': _row,
                    'col': _column,
                    'ws_name': ws.title
                })
                _not_found = False
                break

        if _not_found:
            _flag = False
        _row += 1

    return finded_address